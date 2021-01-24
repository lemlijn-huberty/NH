import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from tkinter import filedialog
import re
import os

def occ_to_site(argument):
  switcher = {
    2158698574: "IFAC",
    2158698673: "IFAC",
    2158698871: "IFAC",
    2158916330: "CHA",
    2161139907: "VDS",
    2161538595: "CUP",
    2166137187: "CSL",
    2174770088: "CSL",
    2174770583: "STA",
    2211863284: "HP",
    2211863482: "MSP",
    2211863581: "STO",
    2253856564: "LBV",
    2256121614: "CUP",
    2287109946: "CUP",
  }
  return switcher.get(argument, "n/a")

# last changed 23/01/2020
# only > XLSX < files

dirname = filedialog.askdirectory()                  # ask for directory of xlsx files to be processed

for filename in os.listdir(dirname):
    if filename.endswith(".xlsx"):
      filename = dirname + '/' + filename
      print(filename)

      workbook = openpyxl.load_workbook(filename)
      try:  # deletes results sheet 'DMFA_occupation' if it already exists (from previous run)
        del workbook['DMFA_occupation']
      except (KeyError, RuntimeError, TypeError, NameError):
        pass

      DMFA = workbook['Occupations']  # Activates selected sheet
      DMFA_TEMP = workbook.copy_worksheet(DMFA)  # copy worksheet to another
      DMFA_TEMP.title = 'DMFA_occTemp'
      DMFA_TEMP.insert_rows(DMFA_TEMP.min_row, 1)  # inserts on top n rows
      DMFA_TEMP['B1'].value = 'Agent'
      DMFA_TEMP['C1'].value = 'NISS'
      DMFA_TEMP['D1'].value = 'numAgent'
      DMFA_TEMP['C2'].value = 'Catégorie Employeur'
      DMFA_TEMP['D2'].value = 'Code travailleur'
      DMFA_TEMP['J3'].value = 'Site'

      df = pd.DataFrame(DMFA_TEMP.values)  # creates dataframe from openpyxl sheet
      df2 = df[0:5]
      na1 = df2.to_numpy().reshape(1, 50)
      li1 = na1[0]
      df4 = pd.DataFrame(columns=li1)
      np.warnings.filterwarnings('ignore', category=np.VisibleDeprecationWarning)
      for ind in df.index:
        if not pd.isnull(df.loc[ind][0]):
          # print(re.findall(r'\d+', df.loc[ind][0])[0])
          df2 = df[ind:ind + 5]
          na2 = df2.to_numpy().reshape(1, 50)
          df3 = pd.DataFrame(na2, columns=li1)
          df4 = df4.append(df3)

      #df.dropna(how='all', inplace=True)  # drop empty rows
      #df.drop(df.tail(1).index, inplace=True)  # drop last n rows
      #df.drop(df.head(1).index, inplace=True)  # drop first n rows

      ws = workbook.create_sheet('DMFA_occupation', 1)

      # inserts dataframe content to new worksheet
      for r in dataframe_to_rows(df4, index=False, header=True):
        ws.append(r)

      for i, row in enumerate(ws):             # go to every row and fills in Cn and Dn with agent name and NISS
          if i != 0:                           # skip header line
            cell = 'A' + str(i + 1)
            cell_agent = 'B' + str(i + 1)
            cell_niss = 'C' + str(i + 1)
            cell_numa = 'D' + str(i + 1)
            if ws[cell].value != None: # extract name and NISS from ex: 'ABBEELS NATHALIE (87090126246) [1]'
                cell_values = ws[cell].value.split('(')
                ws[cell_agent].value = cell_values[0]
                integers = re.findall(r'\d+', cell_values[1])
                ws[cell_niss].value = int(integers[0])
                ws[cell_numa].value = int(integers[1])
            cell = 'L' + str(i + 1)     # extract 952 and 201 from 'Catégorie Employeur : 952 - Code travailleur : 201'
            cell_cat = 'M' + str(i + 1)
            cell_code = 'N' + str(i + 1)
            if ws[cell].value != None:
              integers = re.findall(r'\d+', ws[cell].value)
              ws[cell_cat].value = int(integers[0])
              ws[cell_code].value = int(integers[1])
            cell_nocc = 'AC' + str(i + 1)
            cell_site = 'AD' + str(i + 1)
            ws[cell_site].value = occ_to_site(ws[cell_nocc].value)

      del workbook['DMFA_occTemp']

      sheet = workbook['DMFA_occupation']
      cols = [49, 47, 46, 45, 44, 43, 42, 41, 39, 35, 33, 32, 31, 27, 26, 22, 21, 20, 19, 18, 17, 16, 15, 12, 11,
              10, 9, 8, 7, 6, 5, 1]
      for col in cols:
        sheet.delete_cols(col)

      #col = sheet.column_dimensions['A']
      #col.alignment = Alignment(wrap_text = False, shrink_to_fit = False, indent = 0)
      #col.font = Font(bold=True)
      #Font(name='Arial', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')

                                                 # adjust columns width according to their respective values
      for column_cells in sheet.columns:
          length = max(len(str(cell.value)) for cell in column_cells)
          sheet.column_dimensions[column_cells[0].column_letter].width = length + 3

      workbook.save(filename)



