import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import islice
from tkinter import filedialog
import os

# last changed 23/01/2020
# only > XLSX < files

lookup_values = {'Code : 862': 1, 'Code : 101': 2, 'Code : 201': 3, 'Code : 214': 4, 'Code : 301': 5,
                 'Code : 601': 6, 'Code : 701': 7, 'Code : 845': 8, 'Code : 856': 9, 'Code : 870': 10,
                 'Code : 0001': 11, 'Code : 4001': 12, 'Code : 8000': 13, 'Code : 8003': 14, 'Code : 8004': 15}

def getIndexes(dfObj, value):
    listOfPos = []
    result = dfObj.isin([value])
    seriesObj = result.any()
    columnNames = list(seriesObj[seriesObj == True].index)
    for col in columnNames:
        rows = list(result[col][result[col] == True].index)
        for row in rows:
            listOfPos.append(row)
    return listOfPos

dirname = filedialog.askdirectory()                  # ask for directory of xlsx files to be processed

for filename in os.listdir(dirname):
    if filename.endswith(".xlsx"):

        filename = dirname + '/' + filename
        print(filename)

        workbook = openpyxl.load_workbook(filename)
        try:                                        # deletes results sheet 'NISS' if it already exists (from previous run)
          del workbook['DMFA_modificative']
        except (KeyError, RuntimeError, TypeError, NameError):
          pass

        DMFA = workbook['Feuil1']                       # Activates selected sheet
        DMFA_TEMP = workbook.copy_worksheet(DMFA)       # copy worksheet to another
        DMFA_TEMP.unmerge_cells('A1:G1')
        #DMFA_TEMP.unmerge_cells('B4:C4')
        DMFA_TEMP.title = 'DMFA_ModTemp'
        DMFA_TEMP['A7'].value = 'Type'
        DMFA_TEMP.delete_rows(DMFA_TEMP.min_row, 7)     # deletes first 7 rows
        df = pd.DataFrame(DMFA_TEMP.values)             # creates dataframe from openpyxl sheet
        cols = [0, 5, 6]                                # select columns A, E & F for deletion in next step
        df.drop(df.columns[cols], axis=1, inplace=True) # drop columns (index) list
        df.dropna(how='all', inplace=True)              # drop empty rows
        df.drop(df.tail(1).index, inplace=True)         # drop last n rows
        df.drop(df.head(1).index, inplace=True)         # drop first n rows

        indices = []
        for lookup in lookup_values:
            ListOfPositions = getIndexes(df, lookup)
            try:
                indices.append(ListOfPositions[0])      # exception may occur if one of the codes (lookup_value) are not found
            except:
                pass

        ws = workbook.create_sheet('DMFA_ModTemp_2', 1)
        # inserts dataframe content to new worksheet
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        del workbook['DMFA_ModTemp']

        #workbook.save(filename)
        #workbook = openpyxl.load_workbook(filename)

        sheet = workbook['DMFA_ModTemp_2']

        empty_rows = []
        code = sheet['A1'].value
        for i, row in enumerate(sheet):            # i = 1 up to number of rows in sheet
            cell = 'A' + str(i + 1)                # determines Cell of next row
            if sheet[cell].value != None:          # saves new code
                code = sheet[cell].value
                empty_rows.append(i + 1)
            else:                                  #fills cell value is empty (from last code)
                sheet[cell].value = int(code[7:])  # converts 'Code : 862' to '862'

        for i in sorted(empty_rows, reverse=True): # deletes all rows, previously flagged as empty
            #print(i)                              #   deletion starts with latest row to be deleted
            sheet.delete_rows(i, 1)                #   to avoid reindexing of the remaining rows

        sheet.insert_cols(3, 2)                    # inserts 2 new columns as column 'C' (index 3) and 'D' (index 4)
        for i, row in enumerate(sheet):            # go to every row and fills in Cn and Dn with agent name and NISS
            cell = 'B' + str(i + 1)
            cell_agent = 'C' + str(i + 1)
            cell_niss = 'D' + str(i + 1)
            if sheet[cell].value != None:
                cell_values = sheet[cell].value.split('(')
                sheet[cell_agent].value = cell_values[0]
                cell_values = cell_values[1].split(')')
                sheet[cell_niss].value = int(cell_values[0])
            cell_codet = 'A' + str(i + 1)                     # copy content from column A into column G
            cell_code = 'G' + str(i + 1)
            sheet[cell_code].value = sheet[cell_codet].value

        sheet.insert_rows(1, 1)                    # insert columns header
        sheet['A1'].value = 'CodeTemp'
        sheet['B1'].value = 'Agent'
        sheet['C1'].value = 'Agent Nom'
        sheet['D1'].value = 'NISS'
        sheet['E1'].value = 'Base'
        sheet['F1'].value = 'Montant'
        sheet['G1'].value = 'Code'

        workbook.save(filename)

        fileOcc = filedialog.askopenfilename(initialdir = "c:/",title = "DMFA_occupation",filetypes = (("excel files","*.xlsx"),("all files","*.*")))

        wbOcc = openpyxl.load_workbook(fileOcc)
        occDMFA = wbOcc['DMFA_occupation']          # Activates selected sheet
        data = occDMFA.values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        occDF = pd.DataFrame(data, index=idx, columns=cols)
        #print(occDF)

        modDMFA = workbook['DMFA_ModTemp_2']
        data = modDMFA.values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        modDF = pd.DataFrame(data, index=idx, columns=cols)
        #print(modDF)

        finalDF = modDF.merge(occDF, on='NISS', how='left')

        ws = workbook.create_sheet('DMFA_modificative', 1)
        # inserts dataframe content to new worksheet
        for r in dataframe_to_rows(finalDF, index=False, header=True):
            ws.append(r)

        for column_cells in ws.columns:       # adjust columns width according to their respective values
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 3

        del workbook['DMFA_ModTemp_2']
        workbook.save(filename)
    else:
        continue