import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import filedialog
import os

# last changed 30/12/2020

lookup_values = {'Code : 862': 1, 'Code : 101': 2, 'Code : 201': 3, 'Code : 214': 4, 'Code : 301': 5,
                 'Code : 601': 6, 'Code : 701': 7, 'Code : 856': 8, 'Code : 870': 9, 'Code : 0001': 10,
                 'Code : 4001': 11}

def getIndexes(dfObj, value):
    listOfPos = []
    result = dfObj.isin([value])
    seriesObj = result.any()
    columnNames = list(seriesObj[seriesObj == True].index)
    for col in columnNames:
        rows = list(result[col][result[col] == True].index)
        for row in rows:
            listOfPos.append(row)
    return listOfPos

#filename = filedialog.askopenfilename(initialdir = "c:/",title = "DMFA_récapitulatif",filetypes = (("excel files","*.xls*"),("all files","*.*")))

dirname = filedialog.askdirectory()                  # ask for directory of xlsx files to be processed

for filename in os.listdir(dirname):
    if filename.endswith(".xlsx"):

        filename = dirname + '/' + filename
        print(filename)

        workbook = openpyxl.load_workbook(filename)
        try:                                        # deletes results sheet 'DMFA_NH' if it already exists (from previous run)
          del workbook['DMFA_NH']
        except (KeyError, RuntimeError, TypeError, NameError):
          pass

        DMFA = workbook['Feuil1']                       # Activates selected sheet
        DMFA_TEMP = workbook.copy_worksheet(DMFA)       # copy worksheet to another
        DMFA_TEMP.unmerge_cells('A1:G1')
        DMFA_TEMP.unmerge_cells('B4:C4')
        DMFA_TEMP.title = 'DMFA_Temp'
        DMFA_TEMP['A8'].value = 'Type'
        DMFA_TEMP.delete_rows(DMFA_TEMP.min_row, 7)     # deletes first 7 rows
        df = pd.DataFrame(DMFA_TEMP.values)             # creates dataframe from openpyxl sheet
        cols = [0, 5, 6]
        df.drop(df.columns[cols], axis=1, inplace=True) # drop columns (index) list
        df.dropna(how='all', inplace=True)              # drop empty rows
        df.drop(df.tail(1).index, inplace=True)         # drop last n rows
        df.drop(df.head(1).index, inplace=True)         # drop first n rows

        indices = []
        for lookup in lookup_values:
            ListOfPositions = getIndexes(df, lookup)
            try:
                indices.append(ListOfPositions[0])      # exception may occur if one of the codes (lookup_value) are not found
            except:
                pass

        ws = workbook.create_sheet('DMFA_NH', 1)
        # inserts dataframe content to new worksheet
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

        del workbook['DMFA_Temp']
        workbook.save(filename)

        workbook = openpyxl.load_workbook(filename)
        sheet = workbook['DMFA_NH']

        empty_rows = []
        code = sheet['A1'].value
        for i, row in enumerate(sheet):            # i = 1 up to number of rows in sheet
            cell = 'A' + str(i + 1)                # determines Cell of next row
            if sheet[cell].value != None:          # saves new code
                code = sheet[cell].value
                empty_rows.append(i + 1)
                #sheet[cell].value = None           # clears cell value - to get empty row, to be dropped
                # print(code)
            else:                                  #fills cell value is empty (from last code)
                sheet[cell].value = int(code[7:])  # converts 'Code : 862' to '862'
                # print(cell, sheet[cell].value, code)

        for i in sorted(empty_rows, reverse=True): # deletes all rows, previously flagged as empty
            #print(i)                              #   deletion starts with latest row to be deleted
            sheet.delete_rows(i, 1)                #   to avoid reindexing of the remaining rows

        sheet.insert_cols(3, 2)                    # inserts 2 new columns as column 'C' (index 3) and 'D' (index 4)
        for i, row in enumerate(sheet):            # go to every row and fills in Cn and Dn with agent name and NISS
            cell = 'B' + str(i + 1)
            cell_agent = 'C' + str(i + 1)
            cell_niss = 'D' + str(i + 1)
            if sheet[cell].value != None:
                cell_values = sheet[cell].value.split('(')
                sheet[cell_agent].value = cell_values[0]
                cell_values = cell_values[1].split(')')
                sheet[cell_niss].value = int(cell_values[0])

        sheet.insert_rows(1, 1)                    # insert columns header
        sheet['A1'].value = 'Code'
        sheet['B1'].value = 'Agent'
        sheet['C1'].value = 'Agent Nom'
        sheet['D1'].value = 'Agent NISS'
        sheet['E1'].value = 'Base'
        sheet['F1'].value = 'Montant'

                                                   # adjust columns width according to their respective values
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 3

        sheet.auto_filter.ref = "A:F"              # creates filters on columns A to F

        workbook.save(filename)

    else:
        continue




